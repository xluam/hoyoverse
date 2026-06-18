#!/usr/bin/env python
"""Judge Android UA device performance against a V9 mapping workbook.

Inputs:
  --input CSV/XLSX with a ua/user_agent/userAgent column, or
  --ua one literal user-agent string.

Outputs:
  matched CSV, unmatched detail CSV, unmatched grouped CSV, JSON summary,
  and a Markdown report.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import sys
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Tuple

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

try:
    import openpyxl
except Exception as exc:  # pragma: no cover - environment diagnostic
    raise SystemExit("openpyxl is required. Use the Codex bundled Python runtime.") from exc


MATOMO_MOBILES_URL = (
    "https://raw.githubusercontent.com/matomo-org/device-detector/master/regexes/device/mobiles.yml"
)

STATUS_UNMATCHED = "未命中"
STATUS_CONFLICT = "冲突"
PASS_YES = "是"
PASS_NO = "否"


def norm_exact(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().strip("\"'")
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def norm_product(value) -> str:
    text = norm_exact(value)
    if not text:
        return ""
    text = text.replace("＋", "+")
    for token in ["XIAOMI", "SAMSUNG", "MOBILE", "5G"]:
        text = re.sub(rf"\b{token}\b", "", text)
    return re.sub(r"\s+", " ", text).strip()


def find_latest_v9() -> Optional[Path]:
    candidates_dir = Path.home() / "AppData/Roaming/miHoYo/HoYowave/Shell/File/Downloads"
    candidates = sorted(candidates_dir.glob("CB2_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def choose_record(records: List[dict]) -> Tuple[Optional[dict], str, str]:
    if not records:
        return None, "", ""
    statuses = sorted({str(r["final"]).strip() for r in records if str(r.get("final", "")).strip()})
    if len(statuses) == 1:
        return records[0], statuses[0], ""
    return records[0], f"{STATUS_CONFLICT}:{'/'.join(statuses)}", "/".join(statuses)


def load_v9(v9_path: Path) -> Tuple[dict, dict]:
    wb = openpyxl.load_workbook(v9_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    exact_lookup: Dict[str, List[dict]] = defaultdict(list)
    product_lookup: Dict[str, List[dict]] = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        region = row[0] if len(row) > 0 else None
        device_model = row[1] if len(row) > 1 else None
        product = row[2] if len(row) > 2 else None
        model_id = row[3] if len(row) > 3 else None
        cpu = row[4] if len(row) > 4 else None
        ram = row[5] if len(row) > 5 else None
        final = str(row[10] if len(row) > 10 and row[10] is not None else "").strip()
        if not final:
            continue

        base = {
            "source": "v9_main",
            "region": region,
            "device_model": device_model,
            "product": product,
            "model_id": model_id,
            "cpu": cpu,
            "ram": ram,
            "final": final,
        }

        for col, value in [("model_id", model_id), ("device_model", device_model), ("product", product)]:
            key = norm_exact(value)
            if key:
                rec = dict(base)
                rec["key_col"] = col
                exact_lookup[key].append(rec)

        for col, value in [("device_model", device_model), ("product", product)]:
            key = norm_product(value)
            if key:
                rec = dict(base)
                rec["key_col"] = "loose_" + col
                product_lookup[key].append(rec)

    # Optional patch sheet: B product, D CPU, E performance judgment.
    if len(wb.worksheets) >= 2:
        ws2 = wb.worksheets[1]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            product = row[1] if len(row) > 1 else None
            cpu = row[3] if len(row) > 3 else None
            final = str(row[4] if len(row) > 4 and row[4] is not None else "").strip()
            if not product or not final:
                continue
            rec = {
                "source": "v9_patch",
                "region": "patch",
                "device_model": product,
                "product": product,
                "model_id": "",
                "cpu": cpu,
                "ram": "",
                "final": final,
                "key_col": "patch_product",
            }
            exact_lookup[norm_exact(product)].append(rec)
            product_lookup[norm_product(product)].append(rec)

    return exact_lookup, product_lookup


def yaml_scalar(line: str) -> str:
    value = line.split(":", 1)[1].strip()
    if not value:
        return ""
    if value[0] in ("'", '"'):
        quote = value[0]
        end = len(value) - 1
        while end > 0 and value[end] != quote:
            end -= 1
        body = value[1:end] if end > 0 else value[1:]
        return body.replace("''", "'") if quote == "'" else body.encode("utf-8").decode("unicode_escape")
    return value.split(" #", 1)[0].strip()


def get_matomo_rules_text(rules_path: Optional[Path], cache_path: Path) -> str:
    if rules_path:
        return rules_path.read_text(encoding="utf-8")
    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8")

    cache_path.parent.mkdir(parents=True, exist_ok=True)
    with urllib.request.urlopen(MATOMO_MOBILES_URL, timeout=30) as resp:
        text = resp.read().decode("utf-8")
    cache_path.write_text(text, encoding="utf-8")
    return text


def parse_matomo_mobiles(text: str):
    brand_rules = []
    current = None
    current_model = None

    for raw in text.splitlines():
        line = raw.rstrip("\n")
        if not line.strip() or line.lstrip().startswith("#"):
            continue
        if not line.startswith(" ") and line.endswith(":"):
            brand = line[:-1].strip().strip("\"'")
            current = {"brand": brand, "regex": "", "device": "", "models": []}
            brand_rules.append(current)
            current_model = None
            continue
        if current is None:
            continue

        stripped = line.strip()
        if line.startswith("  ") and not line.startswith("    "):
            if stripped.startswith("regex:"):
                current["regex"] = yaml_scalar(stripped)
            elif stripped.startswith("device:"):
                current["device"] = yaml_scalar(stripped)
            current_model = None
        elif line.startswith("    - "):
            item = stripped[2:].strip()
            current_model = {"regex": "", "model": "", "device": ""}
            current["models"].append(current_model)
            if item.startswith("regex:"):
                current_model["regex"] = yaml_scalar(item)
        elif line.startswith("      ") and current_model is not None:
            if stripped.startswith("regex:"):
                current_model["regex"] = yaml_scalar(stripped)
            elif stripped.startswith("model:"):
                current_model["model"] = yaml_scalar(stripped)
            elif stripped.startswith("device:"):
                current_model["device"] = yaml_scalar(stripped)

    compiled = []
    global_models = []
    for brand_rule in brand_rules:
        try:
            brand_re = re.compile(brand_rule["regex"], re.I) if brand_rule["regex"] else None
        except re.error:
            brand_re = None

        models = []
        for model_rule in brand_rule["models"]:
            try:
                if model_rule["regex"]:
                    model_re = re.compile(model_rule["regex"], re.I)
                    models.append((model_re, model_rule))
                    global_models.append((brand_rule, model_re, model_rule))
            except re.error:
                continue
        compiled.append((brand_rule, brand_re, models))

    return compiled, global_models


def build_by_match(template: str, match: re.Match) -> str:
    out = str(template or "")
    for idx, group in enumerate(match.groups(), start=1):
        out = out.replace(f"${idx}", group or "")
    return out.strip()


def extract_android_model(ua: str) -> str:
    if not isinstance(ua, str) or "Android" not in ua:
        return ""

    match = re.search(r"\(([^)]*Android[^)]*)\)", ua)
    segment = match.group(1) if match else ua
    parts = [part.strip() for part in segment.split(";")]

    for idx, part in enumerate(parts):
        if re.search(r"Android\s+[\w.]+", part, re.I):
            for candidate in parts[idx + 1 :]:
                candidate = candidate.strip()
                if not candidate:
                    continue
                candidate = re.sub(r"\s+Build/.*$", "", candidate, flags=re.I).strip()
                candidate = re.sub(r"\s+wv$", "", candidate, flags=re.I).strip()
                if candidate.lower() in {"wv", "mobile", "tablet", "k"}:
                    continue
                if re.match(r"^[a-z]{2}-[A-Z]{2}$", candidate):
                    continue
                if candidate.startswith("Build/"):
                    continue
                return candidate
    return ""


class DeviceResolver:
    def __init__(self, compiled_rules, global_models):
        self.compiled_rules = compiled_rules
        self.global_models = global_models
        self.cache: Dict[str, dict] = {}

    def resolve(self, raw_model: str, ua: str) -> dict:
        key = raw_model or ua[:120]
        if key in self.cache:
            return self.cache[key]

        targets = [value for value in [raw_model, ua] if value]
        result = {"brand": "", "model": "", "device": "", "method": ""}

        # Brand-first path approximates DeviceDetector's Mobile parser.
        for brand_rule, brand_re, models in self.compiled_rules:
            if brand_re is None or not any(brand_re.search(target) for target in targets):
                continue

            model = ""
            device = brand_rule.get("device") or ""
            for model_re, model_rule in models:
                model_match = None
                for target in targets:
                    model_match = model_re.search(target)
                    if model_match:
                        break
                if model_match:
                    model = build_by_match(model_rule.get("model") or "", model_match)
                    device = model_rule.get("device") or device
                    break
            result = {
                "brand": brand_rule["brand"],
                "model": model,
                "device": device,
                "method": "matomo_brand_then_model",
            }
            break

        # Fallback for PCRE brand regexes Python cannot compile. Only trust
        # model regexes that match from the start of the raw model, otherwise
        # generic codes like X510 can incorrectly win inside SM-X510.
        if not result["model"] and raw_model:
            for brand_rule, model_re, model_rule in self.global_models:
                model_match = model_re.search(raw_model)
                if not model_match or model_match.start() != 0:
                    continue
                result = {
                    "brand": brand_rule["brand"],
                    "model": build_by_match(model_rule.get("model") or "", model_match),
                    "device": model_rule.get("device") or brand_rule.get("device") or "",
                    "method": "matomo_global_model_fallback",
                }
                break

        self.cache[key] = result
        return result


def candidate_products(raw_model: str, detected: dict) -> List[str]:
    out: List[str] = []

    def add(value: str):
        value = re.sub(r"\s+", " ", str(value or "")).strip()
        if value and value not in out:
            out.append(value)

    add(raw_model)
    brand = detected.get("brand") or ""
    model = detected.get("model") or ""
    if not brand or not model:
        return out

    add(f"{brand} {model}")
    add(model)
    brand_upper = brand.upper()
    model_no_5g = re.sub(r"\b5G\b", "", model, flags=re.I).strip()

    if brand_upper == "POCO":
        for prefix in ["Xiaomi Poco", "Xiaomi POCO", "Poco", "POCO"]:
            add(f"{prefix} {model}")
            add(f"{prefix} {model_no_5g}")
    elif brand_upper == "REDMI":
        add(f"Xiaomi Redmi {model}")
        add(f"Redmi {model}")
    elif brand_upper == "SAMSUNG":
        add(f"Samsung {model}")
        add(model)
    elif brand_upper == "GOOGLE":
        add(f"Google {model}")
        add(model)
    elif brand_upper == "XIAOMI":
        add(f"Xiaomi {model}")
        add(model)

    return out


def match_v9(raw_model: str, detected: dict, exact_lookup: dict, product_lookup: dict):
    for candidate in [raw_model, detected.get("model", ""), f"{detected.get('brand', '')} {detected.get('model', '')}"]:
        rec, status, conflict = choose_record(exact_lookup.get(norm_exact(candidate), []))
        if rec:
            return rec, status, "exact:" + rec.get("key_col", ""), candidate, conflict

    for candidate in candidate_products(raw_model, detected):
        rec, status, conflict = choose_record(exact_lookup.get(norm_exact(candidate), []))
        if rec:
            return rec, status, "detector_exact:" + rec.get("key_col", ""), candidate, conflict

    for candidate in candidate_products(raw_model, detected):
        rec, status, conflict = choose_record(product_lookup.get(norm_product(candidate), []))
        if rec:
            return rec, status, "detector_loose:" + rec.get("key_col", ""), candidate, conflict

    return None, STATUS_UNMATCHED, "unmatched", "", ""


def read_ua_rows(input_path: Optional[Path], literal_ua: Optional[str]) -> Iterator[str]:
    if literal_ua:
        yield literal_ua
        return
    if input_path is None:
        raise SystemExit("Provide --input or --ua.")

    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                return
            candidates = ["ua", "user_agent", "userAgent", "User-Agent", "user agent"]
            ua_col = next((col for col in candidates if col in reader.fieldnames), None)
            if ua_col is None:
                ua_col = reader.fieldnames[0]
            for row in reader:
                yield row.get(ua_col, "")
    elif suffix in {".xlsx", ".xlsm"}:
        wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        headers = [str(c or "") for c in next(rows)]
        lower_headers = [h.lower().replace(" ", "_") for h in headers]
        ua_idx = 0
        for candidate in ["ua", "user_agent", "useragent"]:
            if candidate in lower_headers:
                ua_idx = lower_headers.index(candidate)
                break
        for row in rows:
            yield str(row[ua_idx] if len(row) > ua_idx and row[ua_idx] is not None else "")
    else:
        raise SystemExit(f"Unsupported input type: {input_path}")


def write_report(path: Path, summary: dict):
    status = summary["status_counter"]
    rows = summary["rows"]
    lines = [
        "# UA V9 Device Performance Report",
        "",
        f"Generated at: {summary['generated_at']}",
        f"Input rows: {rows}",
        "",
        "## Status Counts",
        "",
    ]
    for key, value in sorted(status.items(), key=lambda kv: kv[1], reverse=True):
        pct = (value / rows * 100) if rows else 0
        lines.append(f"- {key}: {value} ({pct:.2f}%)")

    lines.extend(
        [
            "",
            "## Files",
            "",
            f"- Matched CSV: `{summary['matched_csv']}`",
            f"- Unmatched detail CSV: `{summary['unmatched_detail_csv']}`",
            f"- Unmatched grouped CSV: `{summary['unmatched_grouped_csv']}`",
            "",
            "## Notes",
            "",
            "- `达标` is the only deterministic pass state.",
            "- `未知`, `未命中`, and `冲突:*` need manual review or mapping-table updates.",
            "- UA can be spoofed; use runtime/client hints for high-stakes decisions.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, help="CSV/XLSX containing UA data")
    parser.add_argument("--ua", help="One literal user-agent string")
    parser.add_argument("--v9", type=Path, help="V9 Android model mapping workbook")
    parser.add_argument("--rules", type=Path, help="Optional local Matomo regexes/device/mobiles.yml")
    parser.add_argument("--output-dir", type=Path, default=Path.cwd() / "outputs")
    parser.add_argument("--prefix", default="ua_v9_device_performance")
    args = parser.parse_args(argv)

    v9_path = args.v9 or find_latest_v9()
    if not v9_path or not v9_path.exists():
        raise SystemExit("V9 workbook not found. Pass --v9 path/to/CB2_*.xlsx.")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    rules_cache = args.output_dir / "matomo_mobiles.yml"

    exact_lookup, product_lookup = load_v9(v9_path)
    rules_text = get_matomo_rules_text(args.rules, rules_cache)
    compiled_rules, global_models = parse_matomo_mobiles(rules_text)
    resolver = DeviceResolver(compiled_rules, global_models)

    matched_csv = args.output_dir / f"{args.prefix}_matched.csv"
    unmatched_detail_csv = args.output_dir / f"{args.prefix}_unmatched_detail.csv"
    unmatched_grouped_csv = args.output_dir / f"{args.prefix}_unmatched_grouped.csv"
    summary_json = args.output_dir / f"{args.prefix}_summary.json"
    report_md = args.output_dir / f"{args.prefix}_report.md"

    headers = [
        "ua",
        "ua_android_model",
        "detector_brand",
        "detector_model",
        "detector_device",
        "v9_综合判定",
        "是否达标",
        "match_method",
        "matched_by",
        "v9_region",
        "v9_device_model",
        "v9_product",
        "v9_model_id",
        "v9_cpu",
        "v9_ram",
        "conflict_statuses",
    ]

    status_counter = Counter()
    method_counter = Counter()
    detector_counter = Counter()
    unmatched_groups = {}
    rows = 0

    with matched_csv.open("w", encoding="utf-8-sig", newline="") as matched_handle, unmatched_detail_csv.open(
        "w", encoding="utf-8-sig", newline=""
    ) as unmatched_handle:
        matched_writer = csv.DictWriter(matched_handle, fieldnames=headers)
        unmatched_writer = csv.DictWriter(unmatched_handle, fieldnames=headers)
        matched_writer.writeheader()
        unmatched_writer.writeheader()

        for ua in read_ua_rows(args.input, args.ua):
            rows += 1
            raw_model = extract_android_model(ua)
            detected = resolver.resolve(raw_model, ua)
            rec, status, method, matched_by, conflict = match_v9(raw_model, detected, exact_lookup, product_lookup)

            if detected.get("model"):
                detector_counter["resolved_model"] += 1
            elif detected.get("brand"):
                detector_counter["resolved_brand_only"] += 1
            else:
                detector_counter["not_resolved"] += 1

            pass_flag = PASS_YES if status == "达标" else (PASS_NO if status == "不达标" else "")
            status_counter[status] += 1
            method_counter[method] += 1

            out = {
                "ua": ua,
                "ua_android_model": raw_model,
                "detector_brand": detected.get("brand", ""),
                "detector_model": detected.get("model", ""),
                "detector_device": detected.get("device", ""),
                "v9_综合判定": status,
                "是否达标": pass_flag,
                "match_method": method,
                "matched_by": matched_by,
                "v9_region": rec.get("region", "") if rec else "",
                "v9_device_model": rec.get("device_model", "") if rec else "",
                "v9_product": rec.get("product", "") if rec else "",
                "v9_model_id": rec.get("model_id", "") if rec else "",
                "v9_cpu": rec.get("cpu", "") if rec else "",
                "v9_ram": rec.get("ram", "") if rec else "",
                "conflict_statuses": conflict,
            }
            matched_writer.writerow(out)

            if status == STATUS_UNMATCHED:
                unmatched_writer.writerow(out)
                group_key = (raw_model or "<empty>", detected.get("brand", ""), detected.get("model", ""), detected.get("device", ""))
                if group_key not in unmatched_groups:
                    unmatched_groups[group_key] = {"rows": 0, "sample_ua": ua}
                unmatched_groups[group_key]["rows"] += 1

    with unmatched_grouped_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["ua_android_model", "detector_brand", "detector_model", "detector_device", "rows", "sample_ua"])
        for (model, brand, detected_model, device), data in sorted(
            unmatched_groups.items(), key=lambda item: item[1]["rows"], reverse=True
        ):
            writer.writerow([model, brand, detected_model, device, data["rows"], data["sample_ua"]])

    summary = {
        "input": str(args.input) if args.input else None,
        "v9_xlsx": str(v9_path),
        "matomo_rules": str(args.rules) if args.rules else MATOMO_MOBILES_URL,
        "rows": rows,
        "status_counter": dict(status_counter),
        "method_counter": dict(method_counter),
        "detector_resolution_counter": dict(detector_counter),
        "matched_csv": str(matched_csv),
        "unmatched_detail_csv": str(unmatched_detail_csv),
        "unmatched_grouped_csv": str(unmatched_grouped_csv),
        "report_md": str(report_md),
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
    }
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    write_report(report_md, summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
